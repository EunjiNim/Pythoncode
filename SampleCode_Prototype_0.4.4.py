# Pandas
import openpyxl
import pandas as pd
import numpy as np
import re
import os
import math

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# 기획서 파일 경로
filename = "C:/Users/seaotter/Downloads/Prototype_History0.4.4Data0.3.xlsx"

# 캐릭터 탭 참조 데이터테이블
CharacterExperience = "c:/orario/devqa_release/DataExcel/RatingExperience.xlsx"
CharacterStat = "c:/orario/devqa_release/DataExcel/CharacterStatGroup.xlsx"
CharacterGradeStat = "c:/orario/devqa_release/DataExcel/CharacterGradeStatGroup.xlsx"
CharacterSkillStat = "c:/orario/devqa_release/DataExcel/SkillEffectLevelGroup.xlsx"

# 장비 탭 참조 데이터 테이블
AccessoryRatingExperience = "c:/orario/devqa_release/DataExcel/AccessoryRatingExperience.xlsx"
AccessoryOption = "c:/orario/devqa_release/DataExcel/AccessoryMainOptionGroup.xlsx"
AccessoryName = "c:/orario/devqa_release/DataExcel/ItemAccessory.xlsx"

# 정착가이드 탭 참조 데이터 테이블
MissionReward = "c:/orario/devqa_release/DataExcel/RewardGroup.xlsx"


# 뽑기 탭 참조 데이터 테이블
GachaData = "c:/orario/devqa_release/DataExcel/GachaProbGroup.xlsx"

# 카드 탭 참조 데이터 테이블
AssistCardskill = "c:/orario/devqa_release/DataExcel/SkillEffectLevelGroup.xlsx"

# 완성 파일 경로
output_excel_filename = "C:/Users/seaotter/Downloads/SampleResult_Prototype_History0.4.4.xlsx"


# 데이터 비교 후 셀 배경색 체크해 주는 함수---------------------------------------------------------------------------------------------------------------------
def highlight_diff(row):
    styles = ['background-color: red' if col.startswith('docu_') and row[col] != row[col.replace('docu_', 'data_')] else '' for col in row.index]
    return styles

# 캐릭터 공통 테이블 정보--------------------------------------------------------------------------------------------------------------------------------------
CharactersstatFile = pd.read_excel(filename, sheet_name="캐릭터")

# 캐릭터 숙련도 성장 정보
Character1table = CharactersstatFile[['Unnamed: 3','Unnamed: 4','Unnamed: 5']]

CharacterExperience = pd.read_excel(CharacterExperience, sheet_name="Table")
CharacterExperiencetable = CharacterExperience[['INT', 'INT.1', 'MEMO']]

"""기획서의 데이터와 해당 데이터 문서에 있는 데이터를 서로 합쳐 줌"""
merge_Character1 = pd.merge(Character1table, CharacterExperiencetable
                      , how='left'
                      , left_on='Unnamed: 3'
                      , right_on='INT'
                      , suffixes=('_left', '_right'))

"""공통된 함수를 사용하기 위해 규칙 적용된 컬럼을 적용"""
"""데이터 테이블의 사용되지 않는 행 삭제"""
"""엑셀 테이블에서는 확인할 수 없는 nan 데이터는 정리해 줌"""
merge_Character1.columns = ['docu_1', 'docu_2', 'docu_3', 'data_1', 'data_2', 'data_3']
merge_Character1 = merge_Character1.drop(merge_Character1.index[0:6])
"""의도되지 않은 값이 자꾸 들어가서 모든 행에 데이터가 채워져 있지 않을 경우 행을 삭제함"""
columns_to_check = merge_Character1.columns[merge_Character1.columns.str.match(r'(docu|data)_\d+')]
merge_Character1 = merge_Character1.dropna(subset=columns_to_check, how='any')
merge_Character1 = merge_Character1.reset_index(drop=True)

styled_merge_Character1 = merge_Character1.style.apply(highlight_diff, axis=1)

# 캐릭터 스텟 테이블 비교
Character2table = CharactersstatFile[['Unnamed: 7','Unnamed: 9','Unnamed: 10','Unnamed: 11','Unnamed: 12', 'Unnamed: 13']]

CharacterStats = pd.read_excel(CharacterStat, sheet_name="Table")
CharacterStatstable = CharacterStats[['MEMO.1', 'INT.2', 'INT.3', 'INT.4', 'INT.5', 'INT.6']]

merge_Character2 = pd.merge(Character2table, CharacterStatstable
                      , how='left'
                      , left_on='Unnamed: 7'
                      , right_on='MEMO.1'
                      , suffixes=('_left', '_right'))

merge_Character2.columns = ['docu_1', 'docu_2', 'docu_3', 'docu_4', 'docu_5','docu_6', 'data_1', 'data_2', 'data_3', 'data_4', 'data_5', 'data_6']
merge_Character2 = merge_Character2.drop(merge_Character2.index[0:9])
merge_Character2 = merge_Character2.reset_index(drop=True)

styled_merge_Character2 = merge_Character2.style.apply(highlight_diff, axis=1)



# 캐릭터 초월 스탯
Character3table = CharactersstatFile[['Unnamed: 15','Unnamed: 16','Unnamed: 17','Unnamed: 18','Unnamed: 19', 'Unnamed: 20', 'Unnamed: 21']]

CharacterGradeStats = pd.read_excel(CharacterGradeStat, sheet_name="Table")
CharacterGradeStattable = CharacterGradeStats[['VARCHAR', 'INT.1', 'INT.2', 'INT.3', 'INT.4', 'INT.5', 'INT.6']]

merge_Character3 = pd.merge(Character3table, CharacterGradeStattable
                      , how='left'
                      , left_on='Unnamed: 15'
                      , right_on='VARCHAR'
                      , suffixes=('_left', '_right'))

merge_Character3.columns = ['docu_1', 'docu_2', 'docu_3', 'docu_4', 'docu_5','docu_6','docu_7', 'data_1', 'data_2', 'data_3', 'data_4', 'data_5', 'data_6', 'data_7']
merge_Character3 = merge_Character3.drop(merge_Character2.index[0:9])
merge_Character3 = merge_Character3.dropna(subset=['docu_1', 'data_1'], how='any')
merge_Character3 = merge_Character3.reset_index(drop=True)


styled_merge_Character3 = merge_Character3.style.apply(highlight_diff, axis=1)


# 캐릭터 스킬 계수 
Character4table = CharactersstatFile[['Unnamed: 23','Unnamed: 24','Unnamed: 25']].copy()
Character4table.loc[:, 'SkillType'] = Character4table['Unnamed: 24'].map({'특수기': '스킬1_1', '필살기': '스킬4_1'})

"""NaN 값을 처리하여 문자열로 변경 후 "텍스트1_텍스트2" 형식을 "텍스트2_텍스트1" 형식으로 변경"""
Character4table['Unnamed: 23'] = Character4table['Unnamed: 23'].apply(lambda x: '_'.join(str(x).split('_')[::-1]))

""" 'Unnamed: 25' 열의 데이터를 숫자로 변환, NaN 값을 0으로 대체한 후 반올림 적용"""
Character4table['Unnamed: 25'] = pd.to_numeric(Character4table['Unnamed: 25'], errors='coerce')
Character4table['Unnamed: 25'].fillna(0, inplace=True)
Character4table['Unnamed: 25'] = round(Character4table['Unnamed: 25'] * 100000)

Character4table = Character4table.drop(merge_Character2.index[0:3])
Character4table = Character4table.reset_index(drop=True)

 
CharacterSkillStat = pd.read_excel(CharacterSkillStat, sheet_name="Table")
CharacterSkillStattable = CharacterSkillStat[['VARCHAR.1', 'INT.6']]

"""NaN이 없는 행 추출"""
valid_rows = Character4table.dropna(subset=['Unnamed: 23', 'SkillType'])

"""OR 연산자를 사용하여 여러 문자열을 하나의 정규 표현식으로 합치기"""
pattern = '|'.join(valid_rows[['Unnamed: 23', 'SkillType']].apply(lambda x: f"{x['Unnamed: 23']}_{x['SkillType']}", axis=1).tolist())

"""CharacterSkillStattable의 VARCHAR에 포함되어 있는 행 추출"""
SkillfilterResult = CharacterSkillStattable[CharacterSkillStattable['VARCHAR.1'].str.contains(pattern, na=False)].copy()

"""중복된 데이터 중 첫 번째 데이터만 유지"""
SkillfilterResult = SkillfilterResult.drop_duplicates(subset='VARCHAR.1', keep='first')

"""캐릭터 정보 이름이 너무 간소화되어 있어 동일한 값을 구하기 힘드므로 임시로 생성한 스킬 타입을 붙여줘서 판별에 사용"""
Character4table['Unnamed: 23_SkillType'] = Character4table['Unnamed: 23'] + '_' + Character4table['SkillType']

merge_Character4 = pd.merge(Character4table, SkillfilterResult
                            , how='left'
                            , left_on=['Unnamed: 23_SkillType']
                            , right_on=['VARCHAR.1'])

"""skilltype 관련 컬럼을 지워주고, 데이터 테이블의 캐릭터 정보 이름을 기획서와 동일하게 맞춰 줌"""
merge_Character4 = merge_Character4.drop(['SkillType', 'Unnamed: 23_SkillType'], axis=1)
merge_Character4['VARCHAR.1'] = merge_Character4['VARCHAR.1'].str.replace('(_스킬|스킬).*$', '', regex=True)

merge_Character4.columns = ['docu_1', 'docu_2', 'docu_3', 'data_1', 'data_2']

"""'docu_2' 열에서 NaN 값을 포함하는 행을 삭제"""
merge_Character4 = merge_Character4.dropna(subset=['docu_2'])

def highlight_diff2(row):
    is_diff = (row['docu_3'] != row['data_2']) and (not np.isnan(row['docu_3'])) and (not np.isnan(row['data_2']))
    styles = ['background-color: red' if col == 'docu_3' and is_diff else '' for col in row.index]
    return styles

styled_merge_Character4 = merge_Character4.style.apply(highlight_diff2, axis=1)




# 장비 공통 테이블 정보---------------------------------------------------------------------------------------------------------------------------------------
AccessoryFile = pd.read_excel(filename, sheet_name="장비")


# 장비 성장 정보
Accessory1table = AccessoryFile[['Unnamed: 3','Unnamed: 4','Unnamed: 5', 'Unnamed: 6']]

Accessory1table = Accessory1table.drop(Accessory1table.index[0:13])
Accessory1table = Accessory1table.reset_index(drop=True)


AccessoryRatingExperience = pd.read_excel(AccessoryRatingExperience, sheet_name="Table")
AccessoryRatingExperiencetable = AccessoryRatingExperience[['INT', 'INT.1', 'INT.2', 'MEMO']]

AccessoryRatingExperiencetable = AccessoryRatingExperiencetable.drop(AccessoryRatingExperiencetable.index[0:9])
AccessoryRatingExperiencetable = AccessoryRatingExperiencetable.reset_index(drop=True)


"""merge_Accessory1 = pd.merge(Accessory1table, AccessoryRatingExperiencetable
                      , how='left'
                      , left_on='Unnamed: 3'
                      , right_on='INT'
                      , suffixes=('_left', '_right'))"""

"""키 값 기준이 이상하게 잡혀서 이미 정리된 데이터프레임을 그낭 하나로 바로 머지함"""
merge_Accessory1 = pd.concat([Accessory1table, AccessoryRatingExperiencetable], axis=1)

merge_Accessory1.columns = ['docu_1', 'docu_2', 'docu_3', 'docu_4', 'data_1', 'data_2', 'data_3', 'data_4']
merge_Accessory1 = merge_Accessory1.dropna(subset=['docu_1', 'data_1'], how='any')
merge_Accessory1 = merge_Accessory1.reset_index(drop=True)

styled_merge_Accessory1 = merge_Accessory1.style.apply(highlight_diff, axis=1)



# 장비 별 스텟
Accessory2table = AccessoryFile[['Unnamed: 8','Unnamed: 9','Unnamed: 10', 'Unnamed: 11', 'Unnamed: 12', 'Unnamed: 13', 'Unnamed: 14']]
Accessory2table = Accessory2table.drop(Accessory2table.index[0:3])
Accessory2table = Accessory2table.reset_index(drop=True)

AccessoryOption = pd.read_excel(AccessoryOption, sheet_name="Table")
AccessoryOptiontable = AccessoryOption[['INT', 'VARCHAR', 'INT.1','INT.2', 'INT.3', 'INT.4', 'INT.5', 'INT.6', 'INT.7']]

AccessoryName = pd.read_excel(AccessoryName, sheet_name='Table')
AccessoryNameTable = AccessoryName[['INT', 'VARCHAR']]

"""MainOption 테이블에 아이템 아이디가 없으므로 인덱스 아이디의 마지막 숫자, 숙련도 레벨, 아이템 명을 모두 합쳐 새롭게 만들어 넣어줌"""
"""'INT'를 기준으로 AccessoryOptiontable과 AccessoryNameTable 병합"""
merged_AccessoryOptiontable = pd.merge(AccessoryOptiontable, AccessoryNameTable, on='INT', how='inner')
merged_AccessoryOptiontable = merged_AccessoryOptiontable.drop(merged_AccessoryOptiontable.index[0:13])
merged_AccessoryOptiontable = merged_AccessoryOptiontable.reset_index(drop=True)

merged_AccessoryOptiontable['LastDigit'] = merged_AccessoryOptiontable['INT'].astype(str).apply(lambda x: x[-1])
merged_AccessoryOptiontable['INT'] = merged_AccessoryOptiontable['VARCHAR_y'] + '_' + merged_AccessoryOptiontable['LastDigit'] + '등급 _' + merged_AccessoryOptiontable['INT.1'].astype(str) + '레벨'

"""옵션 설명을 각각 데이터 앞에 설명으로 넣어줌"""
split_value = merged_AccessoryOptiontable['VARCHAR_x'].str.split('/')
split_value = split_value.apply(lambda x: [item.replace('없음', ' - ') for item in x])
merged_AccessoryOptiontable['INT.2'] = split_value.str[0].str.replace(r'\s+', '')
merged_AccessoryOptiontable['INT.4'] = split_value.str[1].str.replace(r'\s+', '')
merged_AccessoryOptiontable['INT.6'] = split_value.str[2].str.replace(r'\s+', '')

"""3번째 옵션 수치는 10만분율로 변환한 후 소수점 2자리수까지만 표시되도록 변환함"""
merged_AccessoryOptiontable['INT.7'] = pd.to_numeric(merged_AccessoryOptiontable['INT.7'], errors='coerce') 
merged_AccessoryOptiontable['INT.7'] = (merged_AccessoryOptiontable['INT.7'] / 100000).round(2)

"""필요없는 컬럼을 지워줌"""
merged_AccessoryOptiontable = merged_AccessoryOptiontable.drop(['VARCHAR_x', 'INT.1', 'VARCHAR_y', 'LastDigit'], axis=1)

merge_Accessory2 = pd.merge(Accessory2table, merged_AccessoryOptiontable
                      , how='left'
                      , left_on='Unnamed: 8'
                      , right_on='INT'
                      , suffixes=('_left', '_right'))

merge_Accessory2.columns = ['docu_1', 'docu_2', 'docu_3', 'docu_4', 'docu_5', 'docu_6', 'docu_7', 'data_1', 'data_2', 'data_3', 'data_4','data_5', 'data_6', 'data_7']
merge_Accessory2 = merge_Accessory2.dropna(subset=['docu_1', 'data_1'], how='any')
merge_Accessory2 = merge_Accessory2.reset_index(drop=True)
merge_Accessory2['docu_4'] = merge_Accessory2['docu_4'].replace(' 없음 ', ' - ')

styled_merge_Accessory2 = merge_Accessory2.style.apply(highlight_diff, axis=1)



# 정착가이드 공통 테이블 정보----------------------------------------------------------------------------------------------------------------------------------
MissionRewardFile = pd.read_excel(filename, sheet_name="정착가이드")

# 정착가이드 보상 데이터
MissionReward1table = MissionRewardFile[['Unnamed: 3','Unnamed: 4','Unnamed: 5','Unnamed: 5']]

MissionReward = pd.read_excel(MissionReward, sheet_name='Table')
MissionRewardtable = MissionReward[['VARCHAR', 'MEMO','INT.6', 'INT.7']]

merge_MissionReward1 = pd.merge(MissionReward1table, MissionRewardtable
                      , how='left'
                      , left_on='Unnamed: 3'
                      , right_on='VARCHAR'
                      , suffixes=('_left', '_right'))

merge_MissionReward1.columns = ['docu_1', 'docu_2', 'docu_3', 'docu_4', 'data_1', 'data_2', 'data_3', 'data_4']
merge_MissionReward1 = merge_MissionReward1.dropna(subset=['docu_1', 'data_1'], how='any')
merge_MissionReward1 = merge_MissionReward1.reset_index(drop=True)

styled_merge_MissionReward1 = merge_MissionReward1.style.apply(highlight_diff, axis=1)







# 뽑기 공통 테이블 정보-----------------------------------------------------------------------------------------------------------------------------------------
GachaFile = pd.read_excel(filename, sheet_name="뽑기")

# 뽑기 보상 데이터
"""소수점으로 표시되어 있는 데이터에 1억을 곱해줌"""
Gacha1table = GachaFile[['Unnamed: 3','Unnamed: 4','Unnamed: 5']]
Gacha1table = Gacha1table.copy()
Gacha1table = Gacha1table.drop(Gacha1table.index[0:3])
Gacha1table = Gacha1table.reset_index(drop=True)
Gacha1table['Unnamed: 5'] = (Gacha1table['Unnamed: 5'] * 100000000).apply(lambda x: round(x, 1)).astype(int)

Gachaprobdata = pd.read_excel(GachaData, sheet_name='Table')
Gachaprobtable = Gachaprobdata[['INT', 'VARCHAR', 'INT.5']]

merge_Gachaprobdata1 = pd.merge(Gacha1table, Gachaprobtable
                      , how='left'
                      , left_on=['Unnamed: 3', 'Unnamed: 4']
                      , right_on=['INT', 'VARCHAR']
                      , suffixes=('_left', '_right'))

merge_Gachaprobdata1.columns = ['docu_1', 'docu_2', 'docu_3', 'data_1', 'data_2', 'data_3']
merge_Gachaprobdata1 = merge_Gachaprobdata1.dropna(subset=['docu_1', 'data_1'], how='any')
merge_Gachaprobdata1 = merge_Gachaprobdata1.reset_index(drop=True)

styled_merge_Gachaprobdata1 = merge_Gachaprobdata1.style.apply(highlight_diff, axis=1)




"""엑셀 데이터 폰트 변경 및 셀 너비 맞춤 함수"""
def style_character_sheet(ws):
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name='맑은 고딕', size=10, color='000000')

    for col in sheet.columns:
        max_length = 0
        column = [cell for cell in col]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width


    return ws


"""계산을 마친 데이터들을 각각의 시트에 별도로 추가해 줌"""
with pd.ExcelWriter(output_excel_filename, engine='openpyxl') as writer:
    styled_merge_Character1.to_excel(writer, index=False, sheet_name='캐릭터_숙련도')
    workbook = writer.book
    sheet = workbook['캐릭터_숙련도']
    style_character_sheet(sheet)

    styled_merge_Character2.to_excel(writer, index=False, sheet_name='캐릭터_성장스탯')
    sheet = workbook['캐릭터_성장스탯']
    style_character_sheet(sheet)

    styled_merge_Character3.to_excel(writer, index=False, sheet_name='캐릭터_초월')
    sheet = workbook['캐릭터_초월']
    style_character_sheet(sheet)
    
    styled_merge_Character4.to_excel(writer, index=False, sheet_name='캐릭터_스킬계수')
    sheet = workbook['캐릭터_스킬계수']
    style_character_sheet(sheet)

    styled_merge_Accessory1.to_excel(writer, index=False, sheet_name='장비_성장')
    sheet = workbook['장비_성장']
    style_character_sheet(sheet)

    styled_merge_Accessory2.to_excel(writer, index=False, sheet_name='장비_장비스탯')
    sheet = workbook['장비_장비스탯']
    style_character_sheet(sheet)

    styled_merge_MissionReward1.to_excel(writer, index=False, sheet_name='정착가이드_보상')
    sheet = workbook['정착가이드_보상']
    style_character_sheet(sheet)

    styled_merge_Gachaprobdata1.to_excel(writer, index=False, sheet_name='뽑기_뽑기데이터')
    sheet = workbook['뽑기_뽑기데이터']
    style_character_sheet(sheet)